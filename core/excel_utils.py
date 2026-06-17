"""Excel utilities - Leitura de planilhas"""

import os
from pathlib import Path
from openpyxl import load_workbook

# Cache simples em memoria
_DATA_CACHE = None
_DATA_CACHE_SIGNATURE = None
_OPERADORES_CACHE = None
_OPERADORES_CACHE_SIGNATURE = None
SHEET_ENV_NAME = "PLANILHA_PROCESSO_SHEET"


def _get_base_dir() -> Path:
    """Retorna a pasta base do projeto"""
    return Path(__file__).resolve().parent.parent


PLANILHAS_DIR = _get_base_dir() / "planilhas"
PLANILHAS_DIR.mkdir(exist_ok=True)

PREFERRED_FILENAMES = [
    "LISTA DE PROCESSO RACK ARAMADO P PILAO.xlsx",
    "LISTA DE PROCESSO RACK ARAMADO P PILÃO.xlsx",
]
EXCEL_PATTERN = "*.xlsx"
OPERADORES_FILENAME = "LISTA DE OPERADORES.xlsx"
REQUIRED_COLUMNS = {"CLIENTE", "ACABADO", "FERRAMENTAL", "PROCESSO"}


def _build_paths_signature(paths: list) -> tuple:
    """Assinatura dos arquivos para detectar mudanças"""
    signature = []
    for path in sorted(paths, key=lambda p: str(p).lower()):
        if not path.exists():
            continue
        stat = path.stat()
        signature.append((str(path.resolve()), stat.st_mtime_ns, stat.st_size))
    return tuple(signature)


def _resolve_excel_paths() -> list:
    """Resolve todas as planilhas disponíveis"""
    search_dirs = [PLANILHAS_DIR, _get_base_dir()]
    found = []

    for folder in search_dirs:
        for name in PREFERRED_FILENAMES:
            candidate = folder / name
            if candidate.exists():
                found.append(candidate)

        matches = sorted(folder.glob(EXCEL_PATTERN))
        for m in matches:
            if m.name == OPERADORES_FILENAME:
                continue
            if m not in found:
                found.append(m)

    if not found:
        raise FileNotFoundError(
            f"Nao encontrei planilhas em {PLANILHAS_DIR} nem em {_get_base_dir()}"
        )

    return found


def _pick_sheet(wb):
    """Seleciona a aba a ser lida"""

    def has_data(rows):
        return any(
            row and any(cell not in (None, "") for cell in row)
            for row in rows
        )

    preferred_name = os.getenv(SHEET_ENV_NAME)
    candidates = []

    if preferred_name and preferred_name in wb.sheetnames:
        candidates.append(preferred_name)

    candidates.append(wb.active.title)

    for name in wb.sheetnames:
        if name not in candidates:
            candidates.append(name)

    for name in candidates:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if has_data(rows):
            return ws, rows

    return wb.active, []


def load_process_data():
    """Lê todas as planilhas e retorna lista de dicionários"""
    global _DATA_CACHE, _DATA_CACHE_SIGNATURE
    excel_paths = _resolve_excel_paths()
    current_signature = _build_paths_signature(excel_paths)

    if _DATA_CACHE is not None and _DATA_CACHE_SIGNATURE == current_signature:
        return _DATA_CACHE

    data = []
    for excel_path in excel_paths:
        caminho_arquivo = os.path.abspath(excel_path)
        wb = load_workbook(caminho_arquivo, data_only=True)
        ws, rows = _pick_sheet(wb)

        if not rows:
            continue

        headers = rows[0]
        if not headers:
            continue
        normalized_headers = {
            str(value).strip().upper()
            for value in headers
            if value not in (None, "")
        }
        if not REQUIRED_COLUMNS.issubset(normalized_headers):
            continue
        for row in rows[1:]:
            if not any(row):
                continue
            item = dict(zip(headers, row))
            data.append(item)

    _DATA_CACHE = data
    _DATA_CACHE_SIGNATURE = current_signature
    return _DATA_CACHE


def get_unique_choices(col_name):
    """Choices simples de uma coluna"""
    data = load_process_data()
    valores = set()

    for row in data:
        value = row.get(col_name)
        if value is None:
            continue
        value = str(value).strip()
        if value:
            valores.add(value)

    choices = [(v, v) for v in sorted(valores)]
    return choices


def get_process_choices_for_ferramental(ferramental):
    """Processos para um ferramental específico"""
    if not ferramental:
        return []

    data = load_process_data()
    valores = set()
    ferramental = str(ferramental).strip()

    for row in data:
        ferr = row.get("FERRAMENTAL")
        proc = row.get("PROCESSO")
        if ferr is None or proc is None:
            continue

        if str(ferr).strip() == ferramental:
            valores.add(str(proc).strip())

    return sorted(valores)


def get_acabados_for_cliente(cliente):
    """Acabados para um cliente específico"""
    if not cliente:
        return []

    data = load_process_data()
    valores = set()
    cliente = str(cliente).strip()

    for row in data:
        cli = row.get("CLIENTE")
        acabado = row.get("ACABADO")
        if cli is None or acabado is None:
            continue

        if str(cli).strip() == cliente:
            valores.add(str(acabado).strip())

    return sorted(valores)


def get_process_choices_for_acabado_e_ferramental(acabado, ferramental):
    """Processos para acabado + ferramental"""
    if not acabado or not ferramental:
        return []

    data = load_process_data()
    valores = set()
    acabado = str(acabado).strip()
    ferramental = str(ferramental).strip()

    for row in data:
        row_acabado = row.get("ACABADO")
        row_ferr = row.get("FERRAMENTAL")
        proc = row.get("PROCESSO")

        if row_acabado is None or row_ferr is None or proc is None:
            continue

        if str(row_acabado).strip() == acabado and str(row_ferr).strip() == ferramental:
            valores.add(str(proc).strip())

    return sorted(valores)


def get_operadores():
    """Retorna lista de operadores"""
    global _OPERADORES_CACHE, _OPERADORES_CACHE_SIGNATURE
    search_paths = [
        PLANILHAS_DIR / OPERADORES_FILENAME,
        _get_base_dir() / OPERADORES_FILENAME,
    ]
    oper_path = next((p for p in search_paths if p.exists()), None)
    if oper_path is None:
        raise FileNotFoundError(f"Nao encontrei {OPERADORES_FILENAME}")

    current_signature = _build_paths_signature([oper_path])
    if _OPERADORES_CACHE is not None and _OPERADORES_CACHE_SIGNATURE == current_signature:
        return _OPERADORES_CACHE

    wb = load_workbook(oper_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        _OPERADORES_CACHE = []
        _OPERADORES_CACHE_SIGNATURE = current_signature
        return _OPERADORES_CACHE

    nomes = []
    start_index = 0
    first_value = rows[0][0] if rows[0] else None
    if isinstance(first_value, str):
        header_value = first_value.strip().upper()
        if "OPERADOR" in header_value or "OPERADORES" in header_value or "NOME" in header_value:
            start_index = 1

    for value, *rest in rows[start_index:]:
        if value is None:
            continue
        nome = str(value).strip()
        if not nome:
            continue
        nomes.append(nome)

    _OPERADORES_CACHE = nomes
    _OPERADORES_CACHE_SIGNATURE = current_signature
    return _OPERADORES_CACHE
