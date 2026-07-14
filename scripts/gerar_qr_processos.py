"""Gera um QR Code por processo cadastrado em uma planilha Excel."""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from pathlib import Path
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit

import qrcode
from openpyxl import load_workbook


COLUNAS_OBRIGATORIAS = (
    "CLIENTE",
    "ACABADO",
    "FERRAMENTAL",
    "PROCESSO",
    "PROCESSO_ID",
)


def normalizar_cabecalho(valor: object) -> str:
    return str(valor or "").strip().upper()


def normalizar_processo_id(valor: object) -> str:
    """Converte o ID para exatamente seis algarismos, preservando zeros."""
    if valor is None:
        raise ValueError("PROCESSO_ID vazio")

    if isinstance(valor, bool):
        raise ValueError(f"PROCESSO_ID inválido: {valor!r}")

    if isinstance(valor, int):
        texto = str(valor)
    elif isinstance(valor, float) and valor.is_integer():
        texto = str(int(valor))
    else:
        texto = str(valor).strip()

    if not texto.isdigit() or len(texto) > 6:
        raise ValueError(
            f"PROCESSO_ID deve conter no máximo seis algarismos: {valor!r}"
        )

    return texto.zfill(6)


def nome_seguro(valor: object, limite: int = 70) -> str:
    texto = unicodedata.normalize("NFKD", str(valor or ""))
    texto = texto.encode("ascii", "ignore").decode("ascii")
    texto = re.sub(r"[^A-Za-z0-9]+", "_", texto).strip("_")
    return (texto[:limite].rstrip("_") or "PROCESSO").upper()


def montar_conteudo_qr(
    processo_id: str,
    modo: str,
    url_base: str | None,
) -> str:
    if modo == "id":
        return processo_id

    if not url_base:
        raise ValueError("--url-base é obrigatória quando --modo=url")

    partes = urlsplit(url_base)
    if partes.scheme not in {"http", "https"} or not partes.netloc:
        raise ValueError("--url-base deve ser uma URL HTTP ou HTTPS válida")

    parametros = dict(parse_qsl(partes.query, keep_blank_values=True))
    parametros["processo_id"] = processo_id
    return urlunsplit(
        (
            partes.scheme,
            partes.netloc,
            partes.path,
            urlencode(parametros),
            partes.fragment,
        )
    )


def carregar_processos(arquivo: Path, aba: str | None = None) -> list[dict]:
    if not arquivo.is_file():
        raise FileNotFoundError(f"Planilha não encontrada: {arquivo}")

    workbook = load_workbook(arquivo, data_only=True, read_only=True)
    try:
        if aba:
            if aba not in workbook.sheetnames:
                raise ValueError(
                    f"Aba {aba!r} não encontrada. Disponíveis: {workbook.sheetnames}"
                )
            worksheet = workbook[aba]
        else:
            worksheet = workbook.active

        linhas = worksheet.iter_rows(values_only=True)
        try:
            cabecalhos_originais = next(linhas)
        except StopIteration as exc:
            raise ValueError("A planilha está vazia") from exc

        cabecalhos = [normalizar_cabecalho(valor) for valor in cabecalhos_originais]
        indices = {nome: indice for indice, nome in enumerate(cabecalhos) if nome}
        faltantes = [nome for nome in COLUNAS_OBRIGATORIAS if nome not in indices]
        if faltantes:
            raise ValueError(f"Colunas obrigatórias ausentes: {', '.join(faltantes)}")

        processos: list[dict] = []
        ids_encontrados: set[str] = set()

        for numero_linha, linha in enumerate(linhas, start=2):
            if not any(valor not in (None, "") for valor in linha):
                continue

            def obter(coluna: str) -> object:
                indice = indices[coluna]
                return linha[indice] if indice < len(linha) else None

            try:
                processo_id = normalizar_processo_id(obter("PROCESSO_ID"))
            except ValueError as exc:
                raise ValueError(f"Linha {numero_linha}: {exc}") from exc

            if processo_id in ids_encontrados:
                raise ValueError(
                    f"Linha {numero_linha}: PROCESSO_ID duplicado: {processo_id}"
                )

            registro = {
                "linha": numero_linha,
                "processo_id": processo_id,
                "cliente": str(obter("CLIENTE") or "").strip(),
                "acabado": str(obter("ACABADO") or "").strip(),
                "ferramental": str(obter("FERRAMENTAL") or "").strip(),
                "processo": str(obter("PROCESSO") or "").strip(),
            }

            campos_vazios = [
                campo
                for campo in ("cliente", "acabado", "ferramental", "processo")
                if not registro[campo]
            ]
            if campos_vazios:
                raise ValueError(
                    f"Linha {numero_linha}: campos vazios: {', '.join(campos_vazios)}"
                )

            ids_encontrados.add(processo_id)
            processos.append(registro)

        if not processos:
            raise ValueError("Nenhum processo válido foi encontrado")

        return processos
    finally:
        workbook.close()


def gerar_qrs(
    arquivo: Path,
    pasta_saida: Path,
    aba: str | None = None,
    modo: str = "id",
    url_base: str | None = None,
) -> list[dict]:
    processos = carregar_processos(arquivo, aba=aba)
    pasta_saida.mkdir(parents=True, exist_ok=True)
    manifesto: list[dict] = []

    for registro in processos:
        conteudo = montar_conteudo_qr(
            registro["processo_id"],
            modo=modo,
            url_base=url_base,
        )
        nome_arquivo = (
            f"{registro['processo_id']}_{nome_seguro(registro['processo'])}.png"
        )
        caminho_qr = pasta_saida / nome_arquivo

        qr = qrcode.QRCode(
            version=None,
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=10,
            border=4,
        )
        qr.add_data(conteudo)
        qr.make(fit=True)
        imagem = qr.make_image(fill_color="black", back_color="white")
        imagem.save(caminho_qr)

        manifesto.append(
            {
                **registro,
                "conteudo_qr": conteudo,
                "arquivo_qr": nome_arquivo,
            }
        )

    caminho_manifesto = pasta_saida / "manifesto_qr.json"
    caminho_manifesto.write_text(
        json.dumps(manifesto, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return manifesto


def construir_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Gera um QR Code PNG para cada PROCESSO_ID da planilha."
    )
    parser.add_argument("--arquivo", required=True, type=Path, help="Planilha .xlsx")
    parser.add_argument("--saida", required=True, type=Path, help="Pasta dos PNGs")
    parser.add_argument("--aba", help="Nome da aba; por padrão usa a aba ativa")
    parser.add_argument(
        "--modo",
        choices=("id", "url"),
        default="id",
        help="Conteúdo do QR: somente o ID ou uma URL com processo_id",
    )
    parser.add_argument(
        "--url-base",
        help="URL do formulário, obrigatória somente no modo url",
    )
    return parser


def main() -> int:
    argumentos = construir_parser().parse_args()
    try:
        manifesto = gerar_qrs(
            arquivo=argumentos.arquivo.resolve(),
            pasta_saida=argumentos.saida.resolve(),
            aba=argumentos.aba,
            modo=argumentos.modo,
            url_base=argumentos.url_base,
        )
    except (FileNotFoundError, OSError, ValueError) as exc:
        print(f"Erro: {exc}")
        return 1

    print(f"{len(manifesto)} QR Codes gerados em: {argumentos.saida.resolve()}")
    print(f"Primeiro ID: {manifesto[0]['processo_id']}")
    print(f"Último ID: {manifesto[-1]['processo_id']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
