"""Processa novas listas de processo e gera IDs/QR Codes com segurança.

O programa foi desenhado para ser executado uma vez por chamada. No Windows, o
Agendador de Tarefas o chama periodicamente; não há um processo residente ocupando
memória enquanto a máquina está ociosa.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import shutil
import sqlite3
import subprocess
import sys
import tempfile
import uuid
from copy import copy
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

import qrcode
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

try:
    from scripts.gerar_qr_base import carregar_base
    from scripts.gerar_qr_processos import (
        montar_conteudo_qr,
        nome_seguro,
        normalizar_processo_id,
    )
except ModuleNotFoundError:
    from gerar_qr_base import carregar_base
    from gerar_qr_processos import (
        montar_conteudo_qr,
        nome_seguro,
        normalizar_processo_id,
    )


COLUNAS_OBRIGATORIAS = ("CLIENTE", "ACABADO", "FERRAMENTAL", "PROCESSO")
TIPOS_VALIDOS = ("producao", "pintura")


class ErroValidacao(ValueError):
    """Erro esperado em uma planilha recebida."""


@dataclass(frozen=True)
class Destino:
    tipo: str
    entrada: Path
    base: Path


@dataclass(frozen=True)
class Configuracao:
    raiz_projeto: Path
    destinos: tuple[Destino, ...]
    pasta_qrs: Path
    pasta_processados: Path
    pasta_rejeitados: Path
    pasta_backups: Path
    pasta_logs: Path
    pasta_estado: Path
    modo_qr: str
    url_base: str | None
    estabilidade_segundos: int
    sincronizar_github: bool
    branch_github: str


@dataclass
class AnalisePlanilha:
    arquivo: Path
    aba: str
    coluna_id: int
    linhas: list[int]
    ids_existentes: dict[int, str]
    registros: dict[int, dict[str, str]]

    @property
    def linhas_sem_id(self) -> list[int]:
        return [linha for linha in self.linhas if linha not in self.ids_existentes]


@dataclass
class Resultado:
    arquivo: str
    tipo: str
    status: str
    mensagem: str
    ids: list[str]
    arquivos_git: list[Path]


def _expandir_caminho(valor: str, raiz: Path) -> Path:
    expandido = os.path.expandvars(os.path.expanduser(valor))
    caminho = Path(expandido)
    return caminho.resolve() if caminho.is_absolute() else (raiz / caminho).resolve()


def carregar_configuracao(caminho: Path) -> Configuracao:
    if not caminho.is_file():
        raise FileNotFoundError(f"Configuração não encontrada: {caminho}")
    dados = json.loads(caminho.read_text(encoding="utf-8"))
    raiz = caminho.parent.parent.resolve()
    raiz = _expandir_caminho(dados.get("raiz_projeto", str(raiz)), caminho.parent)

    pastas = dados["pastas"]
    destinos = tuple(
        Destino(
            tipo=tipo,
            entrada=_expandir_caminho(pastas[f"entrada_{tipo}"], raiz),
            base=_expandir_caminho(pastas[f"base_{tipo}"], raiz),
        )
        for tipo in TIPOS_VALIDOS
    )
    modo = str(dados.get("qr", {}).get("modo", "id")).lower()
    if modo not in {"id", "url"}:
        raise ValueError("qr.modo deve ser 'id' ou 'url'")

    return Configuracao(
        raiz_projeto=raiz,
        destinos=destinos,
        pasta_qrs=_expandir_caminho(pastas["qrs"], raiz),
        pasta_processados=_expandir_caminho(pastas["processados"], raiz),
        pasta_rejeitados=_expandir_caminho(pastas["rejeitados"], raiz),
        pasta_backups=_expandir_caminho(pastas["backups"], raiz),
        pasta_logs=_expandir_caminho(pastas["logs"], raiz),
        pasta_estado=_expandir_caminho(pastas["estado"], raiz),
        modo_qr=modo,
        url_base=dados.get("qr", {}).get("url_base"),
        estabilidade_segundos=max(0, int(dados.get("estabilidade_segundos", 10))),
        sincronizar_github=bool(dados.get("github", {}).get("sincronizar", False)),
        branch_github=str(dados.get("github", {}).get("branch", "main")),
    )


def preparar_pastas(config: Configuracao) -> None:
    for destino in config.destinos:
        destino.entrada.mkdir(parents=True, exist_ok=True)
        destino.base.mkdir(parents=True, exist_ok=True)
    for pasta in (
        config.pasta_qrs,
        config.pasta_processados,
        config.pasta_rejeitados,
        config.pasta_backups,
        config.pasta_logs,
        config.pasta_estado,
    ):
        pasta.mkdir(parents=True, exist_ok=True)


def _agora_iso() -> str:
    return datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")


def registrar_log(config: Configuracao, evento: dict) -> None:
    caminho = config.pasta_logs / f"automacao-{datetime.now():%Y-%m}.jsonl"
    registro = {"data_hora": _agora_iso(), **evento}
    with caminho.open("a", encoding="utf-8") as arquivo:
        arquivo.write(json.dumps(registro, ensure_ascii=False) + "\n")


class TravaExecucao:
    def __init__(self, caminho: Path):
        self.caminho = caminho
        self.descritor: int | None = None

    def __enter__(self):
        try:
            self.descritor = os.open(
                self.caminho,
                os.O_CREAT | os.O_EXCL | os.O_WRONLY,
            )
        except FileExistsError as exc:
            raise RuntimeError("A automação já está sendo executada") from exc
        os.write(self.descritor, f"pid={os.getpid()} data={_agora_iso()}".encode())
        return self

    def __exit__(self, *_):
        if self.descritor is not None:
            os.close(self.descritor)
        self.caminho.unlink(missing_ok=True)


def _hash_arquivo(caminho: Path) -> str:
    digest = hashlib.sha256()
    with caminho.open("rb") as arquivo:
        for bloco in iter(lambda: arquivo.read(1024 * 1024), b""):
            digest.update(bloco)
    return digest.hexdigest()


def _abrir_banco(config: Configuracao) -> sqlite3.Connection:
    conexao = sqlite3.connect(config.pasta_estado / "automacao.sqlite3")
    conexao.execute("PRAGMA journal_mode=WAL")
    conexao.execute(
        """
        CREATE TABLE IF NOT EXISTS sequencia (
            nome TEXT PRIMARY KEY,
            proximo_id INTEGER NOT NULL
        )
        """
    )
    conexao.execute(
        """
        CREATE TABLE IF NOT EXISTS execucoes (
            hash_arquivo TEXT PRIMARY KEY,
            arquivo TEXT NOT NULL,
            tipo TEXT NOT NULL,
            data_hora TEXT NOT NULL,
            primeiro_id TEXT,
            ultimo_id TEXT,
            quantidade INTEGER NOT NULL
        )
        """
    )
    conexao.commit()
    return conexao


def reservar_ids(
    conexao: sqlite3.Connection,
    quantidade: int,
    maior_id_base: int,
) -> list[str]:
    if quantidade <= 0:
        return []
    conexao.execute("BEGIN IMMEDIATE")
    try:
        linha = conexao.execute(
            "SELECT proximo_id FROM sequencia WHERE nome = 'processo'"
        ).fetchone()
        proximo = max(maior_id_base + 1, int(linha[0]) if linha else 1)
        ultimo = proximo + quantidade - 1
        if ultimo > 999999:
            raise RuntimeError("A sequência de seis dígitos foi esgotada")
        conexao.execute(
            """
            INSERT INTO sequencia(nome, proximo_id) VALUES('processo', ?)
            ON CONFLICT(nome) DO UPDATE SET proximo_id = excluded.proximo_id
            """,
            (ultimo + 1,),
        )
        conexao.commit()
    except Exception:
        conexao.rollback()
        raise
    return [str(numero).zfill(6) for numero in range(proximo, ultimo + 1)]


def _normalizar_cabecalho(valor: object) -> str:
    return str(valor or "").strip().upper()


def analisar_planilha(arquivo: Path) -> AnalisePlanilha:
    try:
        workbook = load_workbook(arquivo, data_only=False, read_only=True)
    except Exception as exc:
        raise ErroValidacao(f"não foi possível abrir a planilha: {exc}") from exc
    try:
        encontrada = None
        for worksheet in workbook.worksheets:
            cabecalhos = {
                _normalizar_cabecalho(celula.value): celula.column
                for celula in worksheet[1]
                if _normalizar_cabecalho(celula.value)
            }
            if all(coluna in cabecalhos for coluna in COLUNAS_OBRIGATORIAS):
                encontrada = (worksheet, cabecalhos)
                break
        if encontrada is None:
            raise ErroValidacao(
                "nenhuma aba contém as colunas " + ", ".join(COLUNAS_OBRIGATORIAS)
            )

        worksheet, cabecalhos = encontrada
        coluna_id = cabecalhos.get("PROCESSO_ID", 0)
        if not coluna_id:
            coluna_id = 7
            while worksheet.cell(1, coluna_id).value not in (None, ""):
                coluna_id += 1

        linhas: list[int] = []
        ids_existentes: dict[int, str] = {}
        registros: dict[int, dict[str, str]] = {}
        ids_na_planilha: dict[str, int] = {}

        for numero_linha in range(2, worksheet.max_row + 1):
            valores = {
                coluna: str(worksheet.cell(numero_linha, cabecalhos[coluna]).value or "").strip()
                for coluna in COLUNAS_OBRIGATORIAS
            }
            preenchidos = [coluna for coluna, valor in valores.items() if valor]
            if not preenchidos:
                continue
            faltantes = [coluna for coluna, valor in valores.items() if not valor]
            if faltantes:
                raise ErroValidacao(
                    f"linha {numero_linha}: campos obrigatórios vazios: "
                    + ", ".join(faltantes)
                )

            linhas.append(numero_linha)
            registros[numero_linha] = {
                "cliente": valores["CLIENTE"],
                "acabado": valores["ACABADO"],
                "ferramental": valores["FERRAMENTAL"],
                "processo": valores["PROCESSO"],
            }
            valor_id = worksheet.cell(numero_linha, coluna_id).value
            if valor_id not in (None, ""):
                try:
                    processo_id = normalizar_processo_id(valor_id)
                except ValueError as exc:
                    raise ErroValidacao(f"linha {numero_linha}: {exc}") from exc
                if processo_id in ids_na_planilha:
                    raise ErroValidacao(
                        f"PROCESSO_ID {processo_id} duplicado nas linhas "
                        f"{ids_na_planilha[processo_id]} e {numero_linha}"
                    )
                ids_na_planilha[processo_id] = numero_linha
                ids_existentes[numero_linha] = processo_id

        if not linhas:
            raise ErroValidacao("nenhuma linha de processo preenchida foi encontrada")
        return AnalisePlanilha(
            arquivo=arquivo,
            aba=worksheet.title,
            coluna_id=coluna_id,
            linhas=linhas,
            ids_existentes=ids_existentes,
            registros=registros,
        )
    finally:
        workbook.close()


def _copiar_estilo_coluna(worksheet, origem: int, destino: int, ultima_linha: int) -> None:
    letra_origem = get_column_letter(origem)
    letra_destino = get_column_letter(destino)
    worksheet.column_dimensions[letra_destino].width = (
        worksheet.column_dimensions[letra_origem].width or 14
    )
    for linha in range(1, ultima_linha + 1):
        fonte = worksheet.cell(linha, origem)
        alvo = worksheet.cell(linha, destino)
        if fonte.has_style:
            alvo._style = copy(fonte._style)
        if fonte.number_format:
            alvo.number_format = fonte.number_format
        alvo.alignment = copy(fonte.alignment)
        alvo.protection = copy(fonte.protection)


def preencher_ids(
    origem: Path,
    destino: Path,
    analise: AnalisePlanilha,
    novos_ids: list[str],
) -> dict[int, str]:
    if len(novos_ids) != len(analise.linhas_sem_id):
        raise RuntimeError("quantidade de IDs reservados não corresponde às linhas")
    workbook = load_workbook(origem, data_only=False)
    try:
        worksheet = workbook[analise.aba]
        cabecalho_atual = _normalizar_cabecalho(
            worksheet.cell(1, analise.coluna_id).value
        )
        if cabecalho_atual != "PROCESSO_ID":
            origem_estilo = max(1, analise.coluna_id - 1)
            _copiar_estilo_coluna(
                worksheet,
                origem_estilo,
                analise.coluna_id,
                max(worksheet.max_row, max(analise.linhas)),
            )
        worksheet.cell(1, analise.coluna_id).value = "PROCESSO_ID"
        worksheet.column_dimensions[get_column_letter(analise.coluna_id)].width = 14

        ids_por_linha = dict(analise.ids_existentes)
        for numero_linha, processo_id in zip(analise.linhas_sem_id, novos_ids):
            ids_por_linha[numero_linha] = processo_id
        for numero_linha in analise.linhas:
            celula = worksheet.cell(numero_linha, analise.coluna_id)
            celula.value = ids_por_linha[numero_linha]
            celula.number_format = "@"
            alinhamento = copy(celula.alignment)
            alinhamento.horizontal = "center"
            celula.alignment = alinhamento

        destino.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(destino)
        return ids_por_linha
    finally:
        workbook.close()


def _gerar_qr(conteudo: str, destino: Path) -> None:
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=10,
        border=4,
    )
    qr.add_data(conteudo)
    qr.make(fit=True)
    destino.parent.mkdir(parents=True, exist_ok=True)
    qr.make_image(fill_color="black", back_color="white").save(destino)


def _carregar_manifesto(caminho: Path) -> list[dict]:
    if not caminho.is_file():
        return []
    dados = json.loads(caminho.read_text(encoding="utf-8"))
    if not isinstance(dados, list):
        raise RuntimeError("manifesto_qr.json inválido: esperado uma lista")
    return dados


def _salvar_json_atomico(caminho: Path, dados: object) -> None:
    caminho.parent.mkdir(parents=True, exist_ok=True)
    temporario = caminho.with_name(f".{caminho.name}.{uuid.uuid4().hex}.tmp")
    temporario.write_text(
        json.dumps(dados, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    os.replace(temporario, caminho)


def _nome_disponivel(pasta: Path, nome: str) -> Path:
    destino = pasta / nome
    if not destino.exists():
        return destino
    return pasta / f"{Path(nome).stem}_{datetime.now():%Y%m%d_%H%M%S}{Path(nome).suffix}"


def _arquivo_estavel(arquivo: Path, segundos: int) -> bool:
    if segundos <= 0:
        return True
    idade = datetime.now().timestamp() - arquivo.stat().st_mtime
    return idade >= segundos


def _bases(config: Configuracao) -> list[Path]:
    return [destino.base for destino in config.destinos]


def _ids_da_base(config: Configuracao) -> dict[str, str]:
    registros = carregar_base(_bases(config))
    return {
        registro["processo_id"]: f"{registro['planilha']}, linha {registro['linha']}"
        for registro in registros
    }


def _registrar_execucao(
    conexao: sqlite3.Connection,
    hash_arquivo: str,
    arquivo: str,
    tipo: str,
    ids: list[str],
) -> None:
    conexao.execute(
        """
        INSERT INTO execucoes(
            hash_arquivo, arquivo, tipo, data_hora, primeiro_id, ultimo_id, quantidade
        ) VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            hash_arquivo,
            arquivo,
            tipo,
            _agora_iso(),
            ids[0] if ids else None,
            ids[-1] if ids else None,
            len(ids),
        ),
    )
    conexao.commit()


def processar_arquivo(
    config: Configuracao,
    destino: Destino,
    arquivo: Path,
    conexao: sqlite3.Connection,
    aplicar: bool,
) -> Resultado:
    if arquivo.name.startswith("~$") or arquivo.suffix.lower() != ".xlsx":
        return Resultado(arquivo.name, destino.tipo, "ignorado", "arquivo ignorado", [], [])
    if not _arquivo_estavel(arquivo, config.estabilidade_segundos):
        return Resultado(
            arquivo.name,
            destino.tipo,
            "aguardando",
            "arquivo ainda está sendo copiado ou salvo",
            [],
            [],
        )

    alvo_planilha = destino.base / arquivo.name
    if alvo_planilha.exists():
        raise ErroValidacao(
            f"já existe uma planilha com esse nome na base: {alvo_planilha.name}"
        )
    hash_arquivo = _hash_arquivo(arquivo)
    if conexao.execute(
        "SELECT 1 FROM execucoes WHERE hash_arquivo = ?", (hash_arquivo,)
    ).fetchone():
        raise ErroValidacao("essa mesma planilha já foi processada anteriormente")

    analise = analisar_planilha(arquivo)
    ids_base = _ids_da_base(config)
    for linha, processo_id in analise.ids_existentes.items():
        if processo_id in ids_base:
            raise ErroValidacao(
                f"linha {linha}: PROCESSO_ID {processo_id} já usado em {ids_base[processo_id]}"
            )

    quantidade = len(analise.linhas_sem_id)
    maior_id = max((int(item) for item in ids_base), default=0)
    if not aplicar:
        primeiro = maior_id + 1
        previstos = [str(primeiro + indice).zfill(6) for indice in range(quantidade)]
        return Resultado(
            arquivo.name,
            destino.tipo,
            "simulacao",
            f"planilha válida; {quantidade} IDs seriam atribuídos",
            previstos,
            [],
        )

    novos_ids = reservar_ids(conexao, quantidade, maior_id)
    pasta_execucao = config.pasta_estado / "temporarios" / uuid.uuid4().hex
    planilha_temporaria = pasta_execucao / arquivo.name
    pasta_qr_temporaria = pasta_execucao / "qrs" / arquivo.stem
    manifesto_path = config.pasta_qrs / "manifesto_qr.json"
    qr_final = config.pasta_qrs / arquivo.stem
    publicados: list[Path] = []
    arquivos_git: list[Path] = []
    manifesto_anterior = manifesto_path.read_bytes() if manifesto_path.is_file() else None
    manifesto_atualizado = False

    try:
        ids_por_linha = preencher_ids(
            arquivo, planilha_temporaria, analise, novos_ids
        )
        validada = analisar_planilha(planilha_temporaria)
        if len(validada.ids_existentes) != len(validada.linhas):
            raise RuntimeError("a validação final encontrou linhas sem PROCESSO_ID")

        if qr_final.exists():
            raise ErroValidacao(f"já existe uma pasta de QR para {arquivo.stem}")

        novos_registros: list[dict] = []
        for numero_linha in analise.linhas:
            processo_id = ids_por_linha[numero_linha]
            registro = analise.registros[numero_linha]
            conteudo = montar_conteudo_qr(
                processo_id, config.modo_qr, config.url_base
            )
            nome_qr = f"{processo_id}_{nome_seguro(registro['processo'])}.png"
            _gerar_qr(conteudo, pasta_qr_temporaria / nome_qr)
            novos_registros.append(
                {
                    "linha": numero_linha,
                    "processo_id": processo_id,
                    **registro,
                    "planilha": arquivo.name,
                    "conteudo_qr": conteudo,
                    "pasta_qr": arquivo.stem,
                    "arquivo_qr": f"{arquivo.stem}/{nome_qr}",
                }
            )

        manifesto_atual = _carregar_manifesto(manifesto_path)
        ids_manifesto = {str(item.get("processo_id")) for item in manifesto_atual}
        colisao = ids_manifesto.intersection(item["processo_id"] for item in novos_registros)
        if colisao:
            raise RuntimeError(
                "IDs já existentes no manifesto: " + ", ".join(sorted(colisao))
            )
        manifesto_novo = sorted(
            [*manifesto_atual, *novos_registros],
            key=lambda item: int(item["processo_id"]),
        )

        data = datetime.now().strftime("%Y-%m-%d")
        backup = _nome_disponivel(config.pasta_backups / data / destino.tipo, arquivo.name)
        backup.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(arquivo, backup)

        temporario_alvo = alvo_planilha.with_name(f".{alvo_planilha.name}.tmp")
        shutil.copy2(planilha_temporaria, temporario_alvo)
        os.replace(temporario_alvo, alvo_planilha)
        publicados.append(alvo_planilha)

        shutil.copytree(pasta_qr_temporaria, qr_final)
        publicados.append(qr_final)
        _salvar_json_atomico(manifesto_path, manifesto_novo)
        manifesto_atualizado = True

        processado = _nome_disponivel(
            config.pasta_processados / data / destino.tipo, arquivo.name
        )
        processado.parent.mkdir(parents=True, exist_ok=True)
        shutil.move(str(arquivo), str(processado))

        todos_ids = [ids_por_linha[linha] for linha in analise.linhas]
        _registrar_execucao(
            conexao, hash_arquivo, arquivo.name, destino.tipo, todos_ids
        )
        arquivos_git = [alvo_planilha, manifesto_path, *sorted(qr_final.glob("*.png"))]
        return Resultado(
            arquivo.name,
            destino.tipo,
            "processado",
            f"{len(novos_ids)} IDs criados e {len(todos_ids)} QR Codes gerados",
            todos_ids,
            arquivos_git,
        )
    except Exception:
        if manifesto_atualizado:
            if manifesto_anterior is None:
                manifesto_path.unlink(missing_ok=True)
            else:
                temporario_manifesto = manifesto_path.with_name(
                    f".{manifesto_path.name}.{uuid.uuid4().hex}.rollback"
                )
                temporario_manifesto.write_bytes(manifesto_anterior)
                os.replace(temporario_manifesto, manifesto_path)
        # Somente artefatos novos desta execução podem ser removidos.
        for publicado in reversed(publicados):
            if publicado.is_dir():
                shutil.rmtree(publicado, ignore_errors=True)
            else:
                publicado.unlink(missing_ok=True)
        raise
    finally:
        shutil.rmtree(pasta_execucao, ignore_errors=True)


def rejeitar_arquivo(
    config: Configuracao,
    destino: Destino,
    arquivo: Path,
    mensagem: str,
) -> None:
    data = datetime.now().strftime("%Y-%m-%d")
    rejeitado = _nome_disponivel(
        config.pasta_rejeitados / data / destino.tipo, arquivo.name
    )
    rejeitado.parent.mkdir(parents=True, exist_ok=True)
    shutil.move(str(arquivo), str(rejeitado))
    rejeitado.with_suffix(rejeitado.suffix + ".erro.txt").write_text(
        mensagem, encoding="utf-8"
    )


def _git(config: Configuracao, *argumentos: str) -> subprocess.CompletedProcess:
    return subprocess.run(
        ["git", *argumentos],
        cwd=config.raiz_projeto,
        text=True,
        capture_output=True,
        check=True,
    )


def validar_git_limpo(config: Configuracao) -> None:
    status = _git(config, "status", "--porcelain").stdout.strip()
    if status:
        raise RuntimeError(
            "sincronização GitHub ativada, mas o repositório possui alterações pendentes"
        )
    _git(config, "pull", "--ff-only", "origin", config.branch_github)


def sincronizar_github(config: Configuracao, arquivos: Iterable[Path]) -> None:
    relativos = sorted(
        {
            str(arquivo.resolve().relative_to(config.raiz_projeto)).replace("\\", "/")
            for arquivo in arquivos
        }
    )
    if not relativos:
        return
    _git(config, "add", "--", *relativos)
    _git(
        config,
        "commit",
        "-m",
        f"Automatiza {len(relativos)} arquivo(s) de processos e QR Codes",
    )
    _git(config, "push", "origin", config.branch_github)


def executar(
    config: Configuracao,
    aplicar: bool,
    tipo: str | None = None,
    arquivo_especifico: Path | None = None,
) -> list[Resultado]:
    preparar_pastas(config)
    destinos = [d for d in config.destinos if tipo is None or d.tipo == tipo]
    if not destinos:
        raise ValueError(f"tipo inválido: {tipo}")

    if aplicar and config.sincronizar_github:
        validar_git_limpo(config)

    resultados: list[Resultado] = []
    arquivos_git: list[Path] = []
    with TravaExecucao(config.pasta_estado / "automacao.lock"):
        conexao = _abrir_banco(config)
        try:
            for destino in destinos:
                arquivos = (
                    [arquivo_especifico.resolve()]
                    if arquivo_especifico is not None
                    else sorted(destino.entrada.glob("*.xlsx"))
                )
                for arquivo in arquivos:
                    try:
                        resultado = processar_arquivo(
                            config, destino, arquivo, conexao, aplicar
                        )
                    except (ErroValidacao, OSError, ValueError) as exc:
                        mensagem = str(exc)
                        if aplicar and arquivo.exists():
                            rejeitar_arquivo(config, destino, arquivo, mensagem)
                        resultado = Resultado(
                            arquivo.name,
                            destino.tipo,
                            "rejeitado",
                            mensagem,
                            [],
                            [],
                        )
                    resultados.append(resultado)
                    arquivos_git.extend(resultado.arquivos_git)
                    registrar_log(
                        config,
                        {
                            "arquivo": resultado.arquivo,
                            "tipo": resultado.tipo,
                            "status": resultado.status,
                            "mensagem": resultado.mensagem,
                            "ids": resultado.ids,
                            "modo": "aplicar" if aplicar else "simulacao",
                        },
                    )
        finally:
            conexao.close()

    if aplicar and config.sincronizar_github and arquivos_git:
        sincronizar_github(config, arquivos_git)
    return resultados


def construir_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Inclui PROCESSO_ID em novas planilhas e gera os QR Codes."
    )
    parser.add_argument(
        "--config",
        type=Path,
        default=Path("automacao_qr/config.local.json"),
    )
    parser.add_argument(
        "--aplicar",
        action="store_true",
        help="Efetiva as alterações; sem esta opção executa somente uma simulação.",
    )
    parser.add_argument("--tipo", choices=TIPOS_VALIDOS)
    parser.add_argument(
        "--arquivo",
        type=Path,
        help="Processa uma única planilha; requer também --tipo.",
    )
    return parser


def main() -> int:
    argumentos = construir_parser().parse_args()
    if argumentos.arquivo and not argumentos.tipo:
        print("Erro: --arquivo requer --tipo")
        return 2
    try:
        config = carregar_configuracao(argumentos.config.resolve())
        resultados = executar(
            config,
            aplicar=argumentos.aplicar,
            tipo=argumentos.tipo,
            arquivo_especifico=argumentos.arquivo,
        )
    except Exception as exc:
        print(f"Erro: {exc}")
        return 1

    if not resultados:
        print("Nenhuma planilha nova encontrada.")
        return 0
    for resultado in resultados:
        ids = ""
        if resultado.ids:
            ids = f" | IDs {resultado.ids[0]} a {resultado.ids[-1]}"
        print(
            f"[{resultado.status.upper()}] {resultado.tipo}/{resultado.arquivo}: "
            f"{resultado.mensagem}{ids}"
        )
    return 1 if any(item.status == "rejeitado" for item in resultados) else 0


if __name__ == "__main__":
    raise SystemExit(main())
