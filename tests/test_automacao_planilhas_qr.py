import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from scripts.automacao_planilhas_qr import Configuracao, Destino, executar


class AutomacaoPlanilhasQrTests(unittest.TestCase):
    def setUp(self):
        self.temporario = tempfile.TemporaryDirectory()
        self.raiz = Path(self.temporario.name)
        self.entrada_producao = self.raiz / "entrada" / "producao"
        self.entrada_pintura = self.raiz / "entrada" / "pintura"
        self.base_producao = self.raiz / "planilhas"
        self.base_pintura = self.raiz / "PINTURA"
        for pasta in (
            self.entrada_producao,
            self.entrada_pintura,
            self.base_producao,
            self.base_pintura,
        ):
            pasta.mkdir(parents=True)

        self.config = Configuracao(
            raiz_projeto=self.raiz,
            destinos=(
                Destino("producao", self.entrada_producao, self.base_producao),
                Destino("pintura", self.entrada_pintura, self.base_pintura),
            ),
            pasta_qrs=self.raiz / "qrs",
            pasta_processados=self.raiz / "processados",
            pasta_rejeitados=self.raiz / "rejeitados",
            pasta_backups=self.raiz / "backups",
            pasta_logs=self.raiz / "logs",
            pasta_estado=self.raiz / "estado",
            modo_qr="id",
            url_base=None,
            estabilidade_segundos=0,
            sincronizar_github=False,
            branch_github="main",
        )
        self._criar_planilha(
            self.base_producao / "BASE ATUAL.xlsx",
            linhas=[("Processo anterior", "001142")],
            incluir_id=True,
        )
        self._criar_planilha(
            self.base_pintura / "BASE PINTURA.xlsx",
            linhas=[("Pintura anterior", "001141")],
            incluir_id=True,
        )
        self.config.pasta_qrs.mkdir(parents=True)
        (self.config.pasta_qrs / "manifesto_qr.json").write_text(
            json.dumps(
                [
                    {"processo_id": "001141"},
                    {"processo_id": "001142"},
                ]
            ),
            encoding="utf-8",
        )

    def tearDown(self):
        self.temporario.cleanup()

    def _criar_planilha(self, caminho, linhas, incluir_id=False):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "PROCESSOS"
        cabecalhos = [
            "CLIENTE",
            "ACABADO",
            "FERRAMENTAL",
            "PROCESSO",
            "QNT",
            "QNT TOTAL",
        ]
        if incluir_id:
            cabecalhos.append("PROCESSO_ID")
        worksheet.append(cabecalhos)
        preenchimento = PatternFill("solid", fgColor="5B9BD5")
        for celula in worksheet[1]:
            celula.fill = preenchimento
        for processo, processo_id in linhas:
            valores = ["Cliente", "Produto", "Máquina", processo, 1, "=E2"]
            if incluir_id:
                valores.append(processo_id)
            worksheet.append(valores)
        apoio = workbook.create_sheet("APOIO")
        apoio["A1"] = "CONTEÚDO PRESERVADO"
        apoio["B2"] = "=1+1"
        workbook.save(caminho)

    def test_processa_planilha_preservando_modelo_e_cria_qrs(self):
        entrada = self.entrada_producao / "NOVA LISTA.xlsx"
        self._criar_planilha(
            entrada,
            [("Corte", None), ("Dobra", None)],
            incluir_id=False,
        )

        resultados = executar(self.config, aplicar=True, tipo="producao")

        self.assertEqual(resultados[0].status, "processado")
        self.assertEqual(resultados[0].ids, ["001143", "001144"])
        self.assertFalse(entrada.exists())
        saida = self.base_producao / entrada.name
        workbook = load_workbook(saida, data_only=False)
        try:
            worksheet = workbook["PROCESSOS"]
            self.assertEqual(worksheet["G1"].value, "PROCESSO_ID")
            self.assertEqual(worksheet["G2"].value, "001143")
            self.assertEqual(worksheet["G3"].value, "001144")
            self.assertEqual(worksheet["G2"].number_format, "@")
            self.assertEqual(worksheet["F2"].value, "=E2")
            self.assertEqual(workbook["APOIO"]["A1"].value, "CONTEÚDO PRESERVADO")
            self.assertEqual(worksheet["G1"].fill.fgColor.rgb, worksheet["F1"].fill.fgColor.rgb)
        finally:
            workbook.close()
        qrs = sorted((self.config.pasta_qrs / "NOVA LISTA").glob("*.png"))
        self.assertEqual(len(qrs), 2)
        manifesto = json.loads(
            (self.config.pasta_qrs / "manifesto_qr.json").read_text(encoding="utf-8")
        )
        self.assertEqual(manifesto[-1]["processo_id"], "001144")
        self.assertTrue(list(self.config.pasta_backups.rglob("NOVA LISTA.xlsx")))
        self.assertTrue(list(self.config.pasta_processados.rglob("NOVA LISTA.xlsx")))

    def test_simulacao_nao_altera_planilha_nem_consume_ids(self):
        entrada = self.entrada_producao / "SIMULACAO.xlsx"
        self._criar_planilha(entrada, [("Montagem", None)], incluir_id=False)
        hash_antes = entrada.read_bytes()

        simulacao = executar(self.config, aplicar=False, tipo="producao")
        aplicacao = executar(self.config, aplicar=True, tipo="producao")

        self.assertEqual(simulacao[0].status, "simulacao")
        self.assertEqual(simulacao[0].ids, ["001143"])
        self.assertEqual(hash_antes, list(self.config.pasta_backups.rglob("SIMULACAO.xlsx"))[0].read_bytes())
        self.assertEqual(aplicacao[0].ids, ["001143"])

    def test_rejeita_id_que_ja_existe_na_base(self):
        entrada = self.entrada_producao / "ID REPETIDO.xlsx"
        self._criar_planilha(
            entrada,
            [("Solda", "001142")],
            incluir_id=True,
        )

        resultados = executar(self.config, aplicar=True, tipo="producao")

        self.assertEqual(resultados[0].status, "rejeitado")
        self.assertFalse((self.base_producao / entrada.name).exists())
        self.assertTrue(list(self.config.pasta_rejeitados.rglob("ID REPETIDO.xlsx")))
        erros = list(self.config.pasta_rejeitados.rglob("*.erro.txt"))
        self.assertIn("já usado", erros[0].read_text(encoding="utf-8"))


if __name__ == "__main__":
    unittest.main()
